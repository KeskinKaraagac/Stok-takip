from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

import os
import io
import uuid
import secrets
import logging
import bcrypt
import requests
import jwt as pyjwt
from datetime import datetime, timezone, timedelta, date
from typing import List, Optional, Literal

from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Response, Query, UploadFile, File
from fastapi.responses import StreamingResponse, Response as FastResponse
from fastapi.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr, ConfigDict

from reportlab.lib.pagesizes import A5
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ----- Mongo Setup -----
mongo_url = os.environ["MONGO_URL"]
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ["DB_NAME"]]

# ----- App Setup -----
app = FastAPI(title="Stok-Üretim-Satış API")
api = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
logger = logging.getLogger("stoktakip")

# ----- Constants -----
JWT_ALGORITHM = "HS256"
ROLE_ADMIN = "admin"
ROLE_STAFF = "personel"
ROLE_VIEWER = "rapor"
ALL_ROLES = [ROLE_ADMIN, ROLE_STAFF, ROLE_VIEWER]
WRITE_ROLES = [ROLE_ADMIN, ROLE_STAFF]

# Granular permission keys
ALL_PERMISSIONS = [
    "products.view", "products.create", "products.edit", "products.delete",
    "customers.view", "customers.create", "customers.edit", "customers.delete",
    "productions.view", "productions.create", "productions.delete",
    "sales.view", "sales.create", "sales.delete",
    "stock.view", "stock.adjust",
    "reports.view",
    "users.manage",
]

PERMISSION_GROUPS = [
    {"key": "products", "label": "Ürün Kontrolü", "actions": [
        ("products.view", "Görüntüleme"),
        ("products.create", "Ekleme"),
        ("products.edit", "Düzenleme"),
        ("products.delete", "Silme/Pasif"),
    ]},
    {"key": "customers", "label": "Müşteriler", "actions": [
        ("customers.view", "Görüntüleme"),
        ("customers.create", "Ekleme"),
        ("customers.edit", "Düzenleme"),
        ("customers.delete", "Silme/Pasif"),
    ]},
    {"key": "productions", "label": "Üretim Girişi", "actions": [
        ("productions.view", "Görüntüleme"),
        ("productions.create", "Ekleme"),
        ("productions.delete", "Silme"),
    ]},
    {"key": "sales", "label": "Günlük Satış", "actions": [
        ("sales.view", "Görüntüleme"),
        ("sales.create", "Satış Yapma"),
        ("sales.delete", "Silme/İptal"),
    ]},
    {"key": "stock", "label": "Stok", "actions": [
        ("stock.view", "Görüntüleme"),
        ("stock.adjust", "Manuel Düzeltme"),
    ]},
    {"key": "reports", "label": "Raporlar", "actions": [
        ("reports.view", "Görüntüleme"),
    ]},
    {"key": "users", "label": "Kullanıcılar", "actions": [
        ("users.manage", "Kullanıcı Yönetimi"),
    ]},
]


def default_permissions(role: str):
    if role == ROLE_ADMIN:
        return list(ALL_PERMISSIONS)
    if role == ROLE_VIEWER:
        return ["products.view", "customers.view", "productions.view", "sales.view", "stock.view", "reports.view"]
    if role == ROLE_STAFF:
        return [
            "products.view", "products.create", "products.edit",
            "customers.view", "customers.create", "customers.edit",
            "productions.view", "productions.create",
            "sales.view", "sales.create",
            "stock.view", "stock.adjust",
            "reports.view",
        ]
    return []


def user_permissions(user: dict):
    if user.get("role") == ROLE_ADMIN:
        return list(ALL_PERMISSIONS)
    perms = user.get("permissions")
    if perms is None:
        return default_permissions(user.get("role", ""))
    return perms


# ----- Object Storage Helpers (Emergent) -----
STORAGE_URL = "https://integrations.emergentagent.com/objstore/api/v1/storage"
APP_STORAGE_NAME = os.environ.get("APP_STORAGE_NAME", "stoktakip")
_storage_key: Optional[str] = None


def init_storage() -> Optional[str]:
    global _storage_key
    if _storage_key:
        return _storage_key
    em_key = os.environ.get("EMERGENT_LLM_KEY")
    if not em_key:
        logger.warning("EMERGENT_LLM_KEY missing; storage disabled")
        return None
    try:
        resp = requests.post(f"{STORAGE_URL}/init", json={"emergent_key": em_key}, timeout=30)
        resp.raise_for_status()
        _storage_key = resp.json()["storage_key"]
        logger.info("Object storage initialized")
        return _storage_key
    except Exception as e:
        logger.error(f"Storage init failed: {e}")
        return None


def storage_put(path: str, data: bytes, content_type: str) -> dict:
    key = init_storage()
    if not key:
        raise HTTPException(status_code=500, detail="Storage kullanılamıyor")
    resp = requests.put(
        f"{STORAGE_URL}/objects/{path}",
        headers={"X-Storage-Key": key, "Content-Type": content_type},
        data=data, timeout=120,
    )
    if resp.status_code == 403:
        # key expired -> re-init once
        global _storage_key
        _storage_key = None
        key = init_storage()
        if not key:
            raise HTTPException(status_code=500, detail="Storage yetkisi alınamadı")
        resp = requests.put(
            f"{STORAGE_URL}/objects/{path}",
            headers={"X-Storage-Key": key, "Content-Type": content_type},
            data=data, timeout=120,
        )
    resp.raise_for_status()
    return resp.json()


def storage_get(path: str):
    key = init_storage()
    if not key:
        raise HTTPException(status_code=500, detail="Storage kullanılamıyor")
    resp = requests.get(f"{STORAGE_URL}/objects/{path}", headers={"X-Storage-Key": key}, timeout=60)
    if resp.status_code == 403:
        global _storage_key
        _storage_key = None
        key = init_storage()
        if not key:
            raise HTTPException(status_code=500, detail="Storage yetkisi alınamadı")
        resp = requests.get(f"{STORAGE_URL}/objects/{path}", headers={"X-Storage-Key": key}, timeout=60)
    if resp.status_code == 404:
        return None, None
    resp.raise_for_status()
    return resp.content, resp.headers.get("Content-Type", "application/octet-stream")


# ----- Helpers -----
def now_utc() -> datetime:
    return datetime.now(timezone.utc)


def iso(dt) -> str:
    if isinstance(dt, datetime):
        return dt.isoformat()
    if isinstance(dt, date):
        return datetime(dt.year, dt.month, dt.day, tzinfo=timezone.utc).isoformat()
    return dt


def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(plain: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False


def create_token(sub: str, email: str, role: str, kind: str, minutes: int = 60 * 24 * 7) -> str:
    payload = {
        "sub": sub,
        "email": email,
        "role": role,
        "type": kind,
        "exp": now_utc() + timedelta(minutes=minutes),
    }
    return pyjwt.encode(payload, os.environ["JWT_SECRET"], algorithm=JWT_ALGORITHM)


# ----- Models -----
class UserPublic(BaseModel):
    id: str
    email: EmailStr
    name: str
    role: str
    active: bool = True
    created_at: str


class RegisterIn(BaseModel):
    email: EmailStr
    password: str = Field(min_length=6)
    name: str = Field(min_length=1)
    role: Optional[str] = ROLE_STAFF
    language: Optional[Literal["en", "tr"]] = "en"


class UserCreateIn(BaseModel):
    email: EmailStr
    password: str = Field(min_length=6)
    name: str = Field(min_length=1)
    role: Literal["admin", "personel", "rapor"] = ROLE_STAFF
    language: Literal["en", "tr"] = "en"


class LoginIn(BaseModel):
    email: EmailStr
    password: str


class ProductIn(BaseModel):
    name: str
    code: str
    category: Optional[str] = ""
    unit: Literal["kg", "adet", "koli"] = "adet"
    unit_weight: float = 0.0
    sale_price: float = 0.0
    cost_price: float = 0.0
    current_stock: float = 0.0
    min_stock: float = 0.0
    shelf_life_days: int = 0
    production_date: Optional[str] = None
    expiry_date: Optional[str] = None
    active: bool = True
    notes: Optional[str] = ""


class ProductOut(ProductIn):
    id: str
    created_at: str
    updated_at: str


class ProductionIn(BaseModel):
    date: str
    product_id: str
    quantity: float
    unit: str
    lot_number: Optional[str] = ""
    cost: float = 0.0
    description: Optional[str] = ""


class CustomerIn(BaseModel):
    name: str
    phone: Optional[str] = ""
    email: Optional[str] = ""
    address: Optional[str] = ""
    tax_no: Optional[str] = ""
    customer_type: Optional[str] = "Bireysel"
    active: bool = True
    notes: Optional[str] = ""


class SaleItemIn(BaseModel):
    product_id: str
    quantity: float
    unit_price: float


class SaleIn(BaseModel):
    date: str
    customer_id: str
    items: List[SaleItemIn]
    discount: float = 0.0
    payment_status: Optional[str] = ""
    description: Optional[str] = ""


class StockAdjustIn(BaseModel):
    product_id: str
    movement_type: Literal["manuel", "fire", "iade"]
    quantity: float  # positive = in, negative = out
    description: Optional[str] = ""


# ----- Auth Dependencies -----
async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Oturum açılmamış")
    try:
        payload = pyjwt.decode(token, os.environ["JWT_SECRET"], algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Geçersiz token tipi")
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
        if not user or not user.get("active", True):
            raise HTTPException(status_code=401, detail="Kullanıcı bulunamadı")
        return user
    except pyjwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Oturum süresi doldu")
    except pyjwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Geçersiz token")


def require_roles(*roles):
    async def dep(user: dict = Depends(get_current_user)):
        if user["role"] not in roles:
            raise HTTPException(status_code=403, detail="Yetkisiz erişim")
        return user
    return dep


def require_permission(perm: str):
    async def dep(user: dict = Depends(get_current_user)):
        if user.get("role") == ROLE_ADMIN:
            return user
        perms = user.get("permissions")
        if perms is None:
            perms = default_permissions(user.get("role", ""))
        if perm not in perms:
            raise HTTPException(status_code=403, detail=f"Yetkisiz erişim ({perm})")
        return user
    return dep


# ----- Cookie helper -----
def set_auth_cookies(response: Response, access: str, refresh: str):
    response.set_cookie("access_token", access, httponly=True, secure=False, samesite="lax", max_age=60 * 60 * 24, path="/")
    response.set_cookie("refresh_token", refresh, httponly=True, secure=False, samesite="lax", max_age=60 * 60 * 24 * 7, path="/")


def clear_auth_cookies(response: Response):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")


# ===== AUTH ROUTES =====
@api.post("/auth/register")
async def register(payload: RegisterIn, response: Response):
    email = payload.email.lower().strip()
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=400, detail="Bu e-posta zaten kayıtlı")
    role = payload.role if payload.role in ALL_ROLES else ROLE_STAFF
    user = {
        "id": str(uuid.uuid4()),
        "email": email,
        "name": payload.name,
        "role": role,
        "active": True,
        "language": payload.language or "en",
        "permissions": default_permissions(role),
        "password_hash": hash_password(payload.password),
        "created_at": now_utc().isoformat(),
    }
    await db.users.insert_one(user)
    access = create_token(user["id"], user["email"], user["role"], "access", 60 * 24)
    refresh = create_token(user["id"], user["email"], user["role"], "refresh", 60 * 24 * 7)
    set_auth_cookies(response, access, refresh)
    user.pop("password_hash", None)
    user.pop("_id", None)
    return {"user": user, "access_token": access}


@api.post("/auth/login")
async def login(payload: LoginIn, response: Response):
    email = payload.email.lower().strip()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(payload.password, user.get("password_hash", "")):
        raise HTTPException(status_code=401, detail="E-posta veya şifre hatalı")
    if not user.get("active", True):
        raise HTTPException(status_code=403, detail="Hesap pasif")
    access = create_token(user["id"], user["email"], user["role"], "access", 60 * 24)
    refresh = create_token(user["id"], user["email"], user["role"], "refresh", 60 * 24 * 7)
    set_auth_cookies(response, access, refresh)
    user.pop("password_hash", None)
    user.pop("_id", None)
    return {"user": user, "access_token": access}


@api.post("/auth/logout")
async def logout(response: Response):
    clear_auth_cookies(response)
    return {"ok": True}


@api.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    out = dict(user)
    out["permissions"] = user_permissions(user)
    out["language"] = out.get("language", "en")
    return out


class SelfUpdateIn(BaseModel):
    name: Optional[str] = None
    email: Optional[EmailStr] = None
    language: Optional[Literal["en", "tr"]] = None
    current_password: Optional[str] = None
    new_password: Optional[str] = Field(default=None, min_length=6)

@api.put("/auth/me")
async def update_me(payload: SelfUpdateIn, user: dict = Depends(get_current_user)):
    update = {}
    if payload.name is not None and payload.name.strip():
        update["name"] = payload.name.strip()
    if payload.email is not None:
        em = payload.email.lower().strip()
        clash = await db.users.find_one({"email": em, "id": {"$ne": user["id"]}})
        if clash:
            raise HTTPException(status_code=400, detail="Bu e-posta başka bir kullanıcıya ait")
        update["email"] = em
    if payload.language is not None:
        update["language"] = payload.language
    if payload.new_password:
        full = await db.users.find_one({"id": user["id"]})
        if not full or not verify_password(payload.current_password or "", full.get("password_hash", "")):
            raise HTTPException(status_code=400, detail="Mevcut şifre hatalı")
        update["password_hash"] = hash_password(payload.new_password)
    if update:
        await db.users.update_one({"id": user["id"]}, {"$set": update})
    u = await db.users.find_one({"id": user["id"]}, {"_id": 0, "password_hash": 0})
    if u.get("permissions") is None:
        u["permissions"] = default_permissions(u.get("role", ""))
    return u


@api.get("/permissions")
async def list_permissions(user: dict = Depends(get_current_user)):
    return {"all": ALL_PERMISSIONS, "groups": PERMISSION_GROUPS}


class ForgotPasswordIn(BaseModel):
    email: EmailStr


class ResetPasswordIn(BaseModel):
    token: str
    password: str = Field(min_length=6)


@api.post("/auth/forgot-password")
async def forgot_password(payload: ForgotPasswordIn):
    email = payload.email.lower().strip()
    user = await db.users.find_one({"email": email})
    # Always return success to avoid email enumeration
    if user:
        created_at = now_utc()
        token = secrets.token_urlsafe(32)
        await db.password_reset_tokens.update_many(
            {"user_id": user["id"], "used": False},
            {"$set": {"used": True, "superseded_at": created_at}},
        )
        await db.password_reset_tokens.insert_one({
            "token": token,
            "user_id": user["id"],
            "email": email,
            "expires_at": created_at + timedelta(days=7),
            "used": False,
            "admin_request": True,
            "request_status": "pending",
            "created_at": created_at,
        })
        logger.info(f"[PASSWORD RESET REQUEST] {email} requested an admin password reset")
    return {"ok": True, "message": "Şifre talebiniz yöneticiye iletildi"}


@api.post("/auth/reset-password")
async def reset_password(payload: ResetPasswordIn):
    rec = await db.password_reset_tokens.find_one({"token": payload.token})
    if not rec:
        raise HTTPException(status_code=400, detail="Geçersiz token")
    if rec.get("used"):
        raise HTTPException(status_code=400, detail="Bu bağlantı zaten kullanılmış")
    exp = rec.get("expires_at")
    if isinstance(exp, datetime):
        if exp.tzinfo is None:
            exp = exp.replace(tzinfo=timezone.utc)
        if exp < now_utc():
            raise HTTPException(status_code=400, detail="Bağlantı süresi dolmuş")
    new_hash = hash_password(payload.password)
    await db.users.update_one({"id": rec["user_id"]}, {"$set": {"password_hash": new_hash}})
    await db.password_reset_tokens.update_one({"token": payload.token}, {"$set": {"used": True, "used_at": now_utc()}})
    return {"ok": True, "message": "Şifre başarıyla güncellendi"}


# ===== PRODUCTS =====
@api.get("/products")
async def list_products(active_only: bool = False, user: dict = Depends(get_current_user)):
    q = {"active": True} if active_only else {}
    docs = await db.products.find(q, {"_id": 0}).sort("created_at", -1).to_list(2000)
    return docs


@api.post("/products")
async def create_product(payload: ProductIn, user: dict = Depends(require_permission("products.create"))):
    exists = await db.products.find_one({"code": payload.code})
    if exists:
        raise HTTPException(status_code=400, detail="Bu ürün kodu zaten mevcut")
    now = now_utc().isoformat()
    doc = payload.model_dump()
    doc.update({"id": str(uuid.uuid4()), "created_at": now, "updated_at": now})
    await db.products.insert_one(doc.copy())
    doc.pop("_id", None)
    # Initial stock movement
    if payload.current_stock and payload.current_stock != 0:
        await db.stock_movements.insert_one({
            "id": str(uuid.uuid4()),
            "product_id": doc["id"],
            "date": now,
            "movement_type": "baslangic",
            "in_qty": float(payload.current_stock),
            "out_qty": 0.0,
            "previous_stock": 0.0,
            "new_stock": float(payload.current_stock),
            "reference_id": doc["id"],
            "description": "Başlangıç stoğu",
        })
    return doc


@api.put("/products/{pid}")
async def update_product(pid: str, payload: ProductIn, user: dict = Depends(require_permission("products.edit"))):
    existing = await db.products.find_one({"id": pid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")
    update = payload.model_dump()
    update["updated_at"] = now_utc().isoformat()
    # Stock change via manual edit -> movement
    old_stock = float(existing.get("current_stock", 0))
    new_stock = float(update.get("current_stock", old_stock))
    await db.products.update_one({"id": pid}, {"$set": update})
    if abs(new_stock - old_stock) > 1e-9:
        await db.stock_movements.insert_one({
            "id": str(uuid.uuid4()),
            "product_id": pid,
            "date": now_utc().isoformat(),
            "movement_type": "manuel",
            "in_qty": max(0.0, new_stock - old_stock),
            "out_qty": max(0.0, old_stock - new_stock),
            "previous_stock": old_stock,
            "new_stock": new_stock,
            "reference_id": pid,
            "description": "Ürün düzenleme ile stok güncellemesi",
        })
    updated = await db.products.find_one({"id": pid}, {"_id": 0})
    return updated


@api.delete("/products/{pid}")
async def soft_delete_product(pid: str, user: dict = Depends(require_permission("products.delete"))):
    res = await db.products.update_one({"id": pid}, {"$set": {"active": False, "updated_at": now_utc().isoformat()}})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")
    return {"ok": True}


# ===== PRODUCTIONS =====
@api.get("/productions")
async def list_productions(start: Optional[str] = None, end: Optional[str] = None, user: dict = Depends(get_current_user)):
    q = {}
    if start or end:
        q["date"] = {}
        if start:
            q["date"]["$gte"] = start
        if end:
            q["date"]["$lte"] = end + "T23:59:59"
    docs = await db.productions.find(q, {"_id": 0}).sort("date", -1).to_list(5000)
    return docs


@api.post("/productions")
async def create_production(payload: ProductionIn, user: dict = Depends(require_roles(*WRITE_ROLES))):
    product = await db.products.find_one({"id": payload.product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")
    doc = payload.model_dump()
    doc.update({"id": str(uuid.uuid4()), "created_at": now_utc().isoformat()})
    await db.productions.insert_one(doc.copy())
    doc.pop("_id", None)
    # Update stock
    old = float(product.get("current_stock", 0))
    new = old + float(payload.quantity)
    await db.products.update_one({"id": payload.product_id}, {"$set": {"current_stock": new, "updated_at": now_utc().isoformat()}})
    await db.stock_movements.insert_one({
        "id": str(uuid.uuid4()),
        "product_id": payload.product_id,
        "date": payload.date,
        "movement_type": "uretim",
        "in_qty": float(payload.quantity),
        "out_qty": 0.0,
        "previous_stock": old,
        "new_stock": new,
        "reference_id": doc["id"],
        "description": f"Üretim - Lot: {payload.lot_number or '-'}",
    })
    return doc


@api.delete("/productions/{pid}")
async def delete_production(pid: str, user: dict = Depends(require_permission("productions.delete"))):
    prod = await db.productions.find_one({"id": pid}, {"_id": 0})
    if not prod:
        raise HTTPException(status_code=404, detail="Üretim kaydı bulunamadı")
    # Reverse stock
    product = await db.products.find_one({"id": prod["product_id"]}, {"_id": 0})
    if product:
        old = float(product.get("current_stock", 0))
        new = old - float(prod["quantity"])
        await db.products.update_one({"id": prod["product_id"]}, {"$set": {"current_stock": new}})
        await db.stock_movements.insert_one({
            "id": str(uuid.uuid4()),
            "product_id": prod["product_id"],
            "date": now_utc().isoformat(),
            "movement_type": "iade",
            "in_qty": 0.0,
            "out_qty": float(prod["quantity"]),
            "previous_stock": old,
            "new_stock": new,
            "reference_id": pid,
            "description": "Üretim kaydı silindi",
        })
    await db.productions.delete_one({"id": pid})
    return {"ok": True}


# ===== CUSTOMERS =====
@api.get("/customers")
async def list_customers(user: dict = Depends(get_current_user)):
    docs = await db.customers.find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)
    return docs


@api.get("/customers/{cid}")
async def get_customer(cid: str, user: dict = Depends(get_current_user)):
    c = await db.customers.find_one({"id": cid}, {"_id": 0})
    if not c:
        raise HTTPException(status_code=404, detail="Müşteri bulunamadı")
    sales = await db.sales.find({"customer_id": cid}, {"_id": 0}).sort("date", -1).to_list(2000)
    total = sum(float(s.get("net_total", 0)) for s in sales)
    return {"customer": c, "sales": sales, "total_sales": total}


@api.post("/customers")
async def create_customer(payload: CustomerIn, user: dict = Depends(require_permission("customers.create"))):
    doc = payload.model_dump()
    doc.update({"id": str(uuid.uuid4()), "created_at": now_utc().isoformat()})
    await db.customers.insert_one(doc.copy())
    doc.pop("_id", None)
    return doc


@api.put("/customers/{cid}")
async def update_customer(cid: str, payload: CustomerIn, user: dict = Depends(require_permission("customers.edit"))):
    res = await db.customers.update_one({"id": cid}, {"$set": payload.model_dump()})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Müşteri bulunamadı")
    return await db.customers.find_one({"id": cid}, {"_id": 0})


@api.delete("/customers/{cid}")
async def delete_customer(cid: str, user: dict = Depends(require_permission("customers.delete"))):
    res = await db.customers.update_one({"id": cid}, {"$set": {"active": False}})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Müşteri bulunamadı")
    return {"ok": True}


# ===== SALES =====
@api.get("/sales")
async def list_sales(start: Optional[str] = None, end: Optional[str] = None, customer_id: Optional[str] = None, user: dict = Depends(get_current_user)):
    q = {}
    if start or end:
        q["date"] = {}
        if start:
            q["date"]["$gte"] = start
        if end:
            q["date"]["$lte"] = end + "T23:59:59"
    if customer_id:
        q["customer_id"] = customer_id
    docs = await db.sales.find(q, {"_id": 0}).sort("date", -1).to_list(5000)
    return docs


@api.post("/sales")
async def create_sale(payload: SaleIn, user: dict = Depends(require_permission("sales.create"))):
    if not payload.items:
        raise HTTPException(status_code=400, detail="Satış için en az bir ürün gerekli")
    # Validate stock for each item
    products = {}
    for it in payload.items:
        p = await db.products.find_one({"id": it.product_id}, {"_id": 0})
        if not p:
            raise HTTPException(status_code=404, detail=f"Ürün bulunamadı: {it.product_id}")
        if float(p.get("current_stock", 0)) < float(it.quantity):
            raise HTTPException(status_code=400, detail=f"Yetersiz stok: {p['name']} (mevcut {p['current_stock']})")
        products[it.product_id] = p
    customer = await db.customers.find_one({"id": payload.customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Müşteri bulunamadı")

    items_out = []
    gross = 0.0
    total_weight = 0.0
    for it in payload.items:
        p = products[it.product_id]
        unit_w = float(p.get("unit_weight", 0) or 0)
        unit = (p.get("unit") or "adet").lower()
        # If product unit is kg, the quantity IS the weight; otherwise weight = qty * unit_weight
        item_weight = float(it.quantity) if unit == "kg" else float(it.quantity) * unit_w
        line_total = float(it.quantity) * float(it.unit_price)
        gross += line_total
        total_weight += item_weight
        items_out.append({
            "product_id": it.product_id,
            "product_name": p["name"],
            "product_category": p.get("category", ""),
            "product_unit": p.get("unit", ""),
            "unit_weight": unit_w,
            "quantity": float(it.quantity),
            "weight": item_weight,
            "unit_price": float(it.unit_price),
            "line_total": line_total,
            "cost_price": float(p.get("cost_price", 0)),
        })
    discount = float(payload.discount or 0)
    net = max(0.0, gross - discount)
    sale_id = str(uuid.uuid4())
    sale_doc = {
        "id": sale_id,
        "date": payload.date,
        "customer_id": payload.customer_id,
        "customer_name": customer["name"],
        "items": items_out,
        "gross_total": gross,
        "discount": discount,
        "net_total": net,
        "total_weight": total_weight,
        "total_quantity": sum(float(it.quantity) for it in payload.items),
        "payment_status": payload.payment_status or "",
        "description": payload.description or "",
        "created_at": now_utc().isoformat(),
        "created_by": user["id"],
    }
    await db.sales.insert_one(sale_doc.copy())

    # Apply stock movements
    for it in payload.items:
        p = products[it.product_id]
        old = float(p.get("current_stock", 0))
        new = old - float(it.quantity)
        await db.products.update_one({"id": it.product_id}, {"$set": {"current_stock": new, "updated_at": now_utc().isoformat()}})
        await db.stock_movements.insert_one({
            "id": str(uuid.uuid4()),
            "product_id": it.product_id,
            "date": payload.date,
            "movement_type": "satis",
            "in_qty": 0.0,
            "out_qty": float(it.quantity),
            "previous_stock": old,
            "new_stock": new,
            "reference_id": sale_id,
            "description": f"Satış - {customer['name']}",
        })
    sale_doc.pop("_id", None)
    return sale_doc


@api.get("/sales/{sid}/receipt")
async def sale_receipt_pdf(sid: str, user: dict = Depends(get_current_user)):
    sale = await db.sales.find_one({"id": sid}, {"_id": 0})
    if not sale:
        raise HTTPException(status_code=404, detail="Satış bulunamadı")

    company = await _get_company()
    company_name = company.get("name") or "StokTakip"

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A5, topMargin=12 * mm, bottomMargin=12 * mm, leftMargin=12 * mm, rightMargin=12 * mm)
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontSize=16, textColor=colors.HexColor("#0047AB"), spaceAfter=4)
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#64748b"))

    story = []
    # Header: logo + company name
    logo_bytes_io = None
    if company.get("logo_path"):
        try:
            content, _ = storage_get(company["logo_path"])
            if content:
                logo_bytes_io = io.BytesIO(content)
        except Exception as e:
            logger.warning(f"Logo fetch failed: {e}")
    if logo_bytes_io is None:
        static_path = ROOT_DIR.parent / "frontend" / "public" / "logo.jpg"
        if static_path.exists():
            logo_bytes_io = io.BytesIO(static_path.read_bytes())

    try:
        if logo_bytes_io:
            logo_img = RLImage(logo_bytes_io, width=18 * mm, height=18 * mm)
            header_tbl = Table(
                [[logo_img, Paragraph(f"{company_name}<br/><font size=8 color='#64748b'>SATIŞ FİŞİ</font>", h1)]],
                colWidths=[22 * mm, 90 * mm],
            )
            header_tbl.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.append(header_tbl)
        else:
            story.append(Paragraph(company_name, h1))
            story.append(Paragraph("SATIŞ FİŞİ", small))
    except Exception:
        story.append(Paragraph(company_name, h1))
        story.append(Paragraph("SATIŞ FİŞİ", small))

    # Company contact line
    contact_bits = []
    if company.get("address"):
        contact_bits.append(company["address"])
    if company.get("contact_phone"):
        contact_bits.append("Tel: " + company["contact_phone"])
    if company.get("contact_email"):
        contact_bits.append(company["contact_email"])
    if company.get("tax_no"):
        contact_bits.append("Vergi No: " + company["tax_no"])
    if contact_bits:
        story.append(Paragraph(" • ".join(contact_bits), small))
    story.append(Spacer(1, 8))

    sale_date = (sale.get("date") or "")[:10]
    try:
        y, m, d = sale_date.split("-")
        sale_date_fmt = f"{d}.{m}.{y}"
    except Exception:
        sale_date_fmt = sale_date

    info_data = [
        ["Fiş No:", sale["id"][:8].upper()],
        ["Tarih:", sale_date_fmt],
        ["Müşteri:", sale.get("customer_name", "-")],
        ["Ödeme:", {"odendi": "Ödendi", "bekliyor": "Bekliyor", "kismi": "Kısmi"}.get(sale.get("payment_status"), sale.get("payment_status", "-"))],
    ]
    info_tbl = Table(info_data, colWidths=[28 * mm, 90 * mm])
    info_tbl.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#64748b")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(info_tbl)
    story.append(Spacer(1, 10))

    # Items table
    item_rows = [["Ürün", "Adet", "Birim Fiyat", "Toplam"]]
    for it in sale.get("items", []):
        item_rows.append([
            it.get("product_name", ""),
            f"{float(it.get('quantity', 0)):,.2f}",
            f"£{float(it.get('unit_price', 0)):,.2f}",
            f"£{float(it.get('line_total', 0)):,.2f}",
        ])
    items_tbl = Table(item_rows, colWidths=[58 * mm, 18 * mm, 25 * mm, 25 * mm])
    items_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f1f5f9")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#475569")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e2e8f0")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(items_tbl)
    story.append(Spacer(1, 10))

    # Totals
    totals = [
        ["Brüt Tutar:", f"£{float(sale.get('gross_total', 0)):,.2f}"],
        ["İskonto:", f"£{float(sale.get('discount', 0)):,.2f}"],
        ["NET TUTAR:", f"£{float(sale.get('net_total', 0)):,.2f}"],
    ]
    t_tbl = Table(totals, colWidths=[60 * mm, 35 * mm], hAlign="RIGHT")
    t_tbl.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -2), 10),
        ("FONTSIZE", (0, -1), (-1, -1), 13),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, -1), (-1, -1), colors.HexColor("#0047AB")),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("LINEABOVE", (0, -1), (-1, -1), 1, colors.HexColor("#0f172a")),
        ("TOPPADDING", (0, -1), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(t_tbl)

    if sale.get("description"):
        story.append(Spacer(1, 10))
        story.append(Paragraph(f"<i>Açıklama: {sale['description']}</i>", small))

    story.append(Spacer(1, 16))
    story.append(Paragraph(f"Bu fiş {company_name} tarafından {now_utc().strftime('%d.%m.%Y %H:%M')} UTC tarihinde oluşturulmuştur.", small))

    doc.build(story)
    buf.seek(0)
    filename = f"fis-{sale['id'][:8]}.pdf"
    return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f'inline; filename="{filename}"'})


@api.delete("/sales/{sid}")
async def delete_sale(sid: str, user: dict = Depends(require_permission("sales.delete"))):
    sale = await db.sales.find_one({"id": sid}, {"_id": 0})
    if not sale:
        raise HTTPException(status_code=404, detail="Satış bulunamadı")
    # Reverse stock for each item
    for it in sale.get("items", []):
        p = await db.products.find_one({"id": it["product_id"]}, {"_id": 0})
        if p:
            old = float(p.get("current_stock", 0))
            new = old + float(it["quantity"])
            await db.products.update_one({"id": it["product_id"]}, {"$set": {"current_stock": new}})
            await db.stock_movements.insert_one({
                "id": str(uuid.uuid4()),
                "product_id": it["product_id"],
                "date": now_utc().isoformat(),
                "movement_type": "iade",
                "in_qty": float(it["quantity"]),
                "out_qty": 0.0,
                "previous_stock": old,
                "new_stock": new,
                "reference_id": sid,
                "description": "Satış iptali / iade",
            })
    await db.sales.delete_one({"id": sid})
    return {"ok": True}


# ===== STOCK =====
@api.get("/stock/movements")
async def list_movements(product_id: Optional[str] = None, start: Optional[str] = None, end: Optional[str] = None, user: dict = Depends(get_current_user)):
    q = {}
    if product_id:
        q["product_id"] = product_id
    if start or end:
        q["date"] = {}
        if start:
            q["date"]["$gte"] = start
        if end:
            q["date"]["$lte"] = end + "T23:59:59"
    docs = await db.stock_movements.find(q, {"_id": 0}).sort("date", -1).to_list(5000)
    return docs


@api.post("/stock/adjust")
async def adjust_stock(payload: StockAdjustIn, user: dict = Depends(require_permission("stock.adjust"))):
    product = await db.products.find_one({"id": payload.product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")
    old = float(product.get("current_stock", 0))
    qty = float(payload.quantity)
    new = old + qty
    if new < 0:
        raise HTTPException(status_code=400, detail="Stok negatife düşemez")
    await db.products.update_one({"id": payload.product_id}, {"$set": {"current_stock": new, "updated_at": now_utc().isoformat()}})
    mov_id = str(uuid.uuid4())
    await db.stock_movements.insert_one({
        "id": mov_id,
        "product_id": payload.product_id,
        "date": now_utc().isoformat(),
        "movement_type": payload.movement_type,
        "in_qty": max(0.0, qty),
        "out_qty": max(0.0, -qty),
        "previous_stock": old,
        "new_stock": new,
        "reference_id": mov_id,
        "description": payload.description or "",
    })
    return {"ok": True, "previous_stock": old, "new_stock": new}


# ===== COMPANY PROFILE =====
class CompanyIn(BaseModel):
    name: str = Field(min_length=1, max_length=120)
    contact_phone: Optional[str] = ""
    contact_email: Optional[str] = ""
    address: Optional[str] = ""
    tax_no: Optional[str] = ""
    website: Optional[str] = ""


COMPANY_DOC_ID = "company_profile"


def _empty_company():
    return {
        "id": COMPANY_DOC_ID,
        "name": "StokTakip",
        "contact_phone": "",
        "contact_email": "",
        "address": "",
        "tax_no": "",
        "website": "",
        "logo_path": None,
        "logo_ext": None,
        "logo_updated_at": None,
        "updated_at": now_utc().isoformat(),
    }


async def _get_company():
    doc = await db.company.find_one({"id": COMPANY_DOC_ID}, {"_id": 0})
    if not doc:
        doc = _empty_company()
        await db.company.insert_one(doc.copy())
    return doc


@api.get("/company")
async def get_company():
    """Public — used by login page too."""
    doc = await _get_company()
    return {
        "name": doc.get("name", "StokTakip"),
        "contact_phone": doc.get("contact_phone", ""),
        "contact_email": doc.get("contact_email", ""),
        "address": doc.get("address", ""),
        "tax_no": doc.get("tax_no", ""),
        "website": doc.get("website", ""),
        "has_logo": bool(doc.get("logo_path")),
        "logo_updated_at": doc.get("logo_updated_at"),
        "updated_at": doc.get("updated_at"),
    }


@api.put("/company")
async def update_company(payload: CompanyIn, user: dict = Depends(require_roles(ROLE_ADMIN))):
    update = payload.model_dump()
    update["updated_at"] = now_utc().isoformat()
    await db.company.update_one({"id": COMPANY_DOC_ID}, {"$set": update}, upsert=True)
    return await get_company()


_ALLOWED_LOGO_TYPES = {"image/jpeg", "image/png", "image/webp", "image/gif"}
_EXT_BY_TYPE = {"image/jpeg": "jpg", "image/png": "png", "image/webp": "webp", "image/gif": "gif"}


@api.post("/company/logo")
async def upload_company_logo(file: UploadFile = File(...), user: dict = Depends(require_roles(ROLE_ADMIN))):
    ctype = (file.content_type or "").lower()
    if ctype not in _ALLOWED_LOGO_TYPES:
        raise HTTPException(status_code=400, detail="Sadece JPG, PNG, WEBP veya GIF yüklenebilir")
    data = await file.read()
    if len(data) > 2 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Logo en fazla 2 MB olabilir")
    if len(data) < 32:
        raise HTTPException(status_code=400, detail="Geçersiz dosya")
    ext = _EXT_BY_TYPE.get(ctype, "bin")
    path = f"{APP_STORAGE_NAME}/company/logo-{uuid.uuid4()}.{ext}"
    result = storage_put(path, data, ctype)
    now = now_utc().isoformat()
    await db.company.update_one(
        {"id": COMPANY_DOC_ID},
        {"$set": {"logo_path": result["path"], "logo_ext": ext, "logo_content_type": ctype, "logo_updated_at": now, "updated_at": now}},
        upsert=True,
    )
    return await get_company()


@api.get("/company/logo")
async def get_company_logo(v: Optional[str] = None):
    """Public — serves the uploaded logo so it can be embedded as <img src> without auth."""
    doc = await db.company.find_one({"id": COMPANY_DOC_ID}, {"_id": 0})
    if not doc or not doc.get("logo_path"):
        raise HTTPException(status_code=404, detail="Logo yüklenmemiş")
    content, ctype = storage_get(doc["logo_path"])
    if content is None:
        raise HTTPException(status_code=404, detail="Logo bulunamadı")
    return FastResponse(content=content, media_type=ctype or doc.get("logo_content_type") or "image/jpeg", headers={"Cache-Control": "public, max-age=300"})


@api.delete("/company/logo")
async def remove_company_logo(user: dict = Depends(require_roles(ROLE_ADMIN))):
    now = now_utc().isoformat()
    await db.company.update_one(
        {"id": COMPANY_DOC_ID},
        {"$set": {"logo_path": None, "logo_ext": None, "logo_content_type": None, "logo_updated_at": None, "updated_at": now}},
    )
    return await get_company()


# ===== EXCEL EXPORT =====
HEADER_FILL = PatternFill(start_color="0047AB", end_color="0047AB", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN = Side(border_style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _build_xlsx(title: str, headers, rows, money_cols=(), number_cols=()):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = BORDER
            if c_idx in money_cols:
                cell.number_format = '£#,##0.00'
            elif c_idx in number_cols:
                cell.number_format = '#,##0.00'
    # Autosize approximated
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        max_len = len(str(headers[col - 1]))
        for r in rows:
            v = r[col - 1] if col - 1 < len(r) else ""
            ln = len(str(v)) if v is not None else 0
            if ln > max_len:
                max_len = ln
        ws.column_dimensions[letter].width = min(max_len + 3, 40)
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _xlsx_response(buf: io.BytesIO, filename: str):
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _fmt_date(v):
    if not v:
        return ""
    s = str(v)[:10]
    try:
        y, m, d = s.split("-")
        return f"{d}.{m}.{y}"
    except Exception:
        return s


@api.get("/export/products.xlsx")
async def export_products(user: dict = Depends(require_permission("products.view"))):
    products = await db.products.find({}, {"_id": 0}).sort("name", 1).to_list(5000)
    headers = ["Ad", "Kod", "Kategori", "Birim", "Birim Ağırlık", "Stok", "Min Stok", "Maliyet (£)", "Satış Fiyatı (£)", "Stok Maliyet Değeri (£)", "Stok Satış Değeri (£)", "Üretim Tarihi", "SKT", "Aktif", "Notlar"]
    rows = []
    for p in products:
        stock = float(p.get("current_stock", 0))
        cost = float(p.get("cost_price", 0))
        sale = float(p.get("sale_price", 0))
        rows.append([
            p.get("name", ""), p.get("code", ""), p.get("category", ""), p.get("unit", ""),
            float(p.get("unit_weight", 0)), stock, float(p.get("min_stock", 0)),
            cost, sale, stock * cost, stock * sale,
            _fmt_date(p.get("production_date")), _fmt_date(p.get("expiry_date")),
            "Evet" if p.get("active") else "Hayır", p.get("notes", ""),
        ])
    buf = _build_xlsx("Ürünler", headers, rows, money_cols=(8, 9, 10, 11), number_cols=(5, 6, 7))
    return _xlsx_response(buf, "urunler.xlsx")


@api.get("/export/stock.xlsx")
async def export_stock(user: dict = Depends(require_permission("reports.view"))):
    products = await db.products.find({"active": True}, {"_id": 0}).sort("name", 1).to_list(5000)
    headers = ["Ürün", "Kod", "Kategori", "Birim", "Güncel Stok", "Maliyet (£)", "Satış (£)", "Mali Değer (£)", "Satış Değeri (£)"]
    rows = []
    for p in products:
        stock = float(p.get("current_stock", 0))
        cost = float(p.get("cost_price", 0))
        sale = float(p.get("sale_price", 0))
        rows.append([
            p.get("name", ""), p.get("code", ""), p.get("category", ""), p.get("unit", ""),
            stock, cost, sale, stock * cost, stock * sale,
        ])
    buf = _build_xlsx("Stok Raporu", headers, rows, money_cols=(6, 7, 8, 9), number_cols=(5,))
    return _xlsx_response(buf, "stok-raporu.xlsx")


@api.get("/export/sales.xlsx")
async def export_sales(start: Optional[str] = None, end: Optional[str] = None, user: dict = Depends(require_permission("sales.view"))):
    q = {}
    if start or end:
        q["date"] = {}
        if start:
            q["date"]["$gte"] = start
        if end:
            q["date"]["$lte"] = end + "T23:59:59"
    sales = await db.sales.find(q, {"_id": 0}).sort("date", -1).to_list(10000)
    headers = ["Tarih", "Fiş No", "Müşteri", "Ürünler", "Brüt (£)", "İskonto (£)", "Net (£)", "Ödeme Durumu", "Açıklama"]
    rows = []
    status_map = {"odendi": "Ödendi", "bekliyor": "Bekliyor", "kismi": "Kısmi"}
    for s in sales:
        items_str = "; ".join(f"{it.get('product_name','')} ×{float(it.get('quantity',0)):.2f}" for it in s.get("items", []))
        rows.append([
            _fmt_date(s.get("date")),
            s.get("id", "")[:8].upper(),
            s.get("customer_name", ""),
            items_str,
            float(s.get("gross_total", 0)),
            float(s.get("discount", 0)),
            float(s.get("net_total", 0)),
            status_map.get(s.get("payment_status"), s.get("payment_status", "")),
            s.get("description", ""),
        ])
    buf = _build_xlsx("Satışlar", headers, rows, money_cols=(5, 6, 7))
    return _xlsx_response(buf, "satislar.xlsx")


@api.get("/export/productions.xlsx")
async def export_productions(start: Optional[str] = None, end: Optional[str] = None, user: dict = Depends(require_permission("productions.view"))):
    q = {}
    if start or end:
        q["date"] = {}
        if start:
            q["date"]["$gte"] = start
        if end:
            q["date"]["$lte"] = end + "T23:59:59"
    prods = await db.productions.find(q, {"_id": 0}).sort("date", -1).to_list(10000)
    pmap = {pp["id"]: pp["name"] for pp in await db.products.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(5000)}
    headers = ["Tarih", "Ürün", "Miktar", "Birim", "Lot/Parti No", "Maliyet (£)", "Açıklama"]
    rows = []
    for p in prods:
        rows.append([
            _fmt_date(p.get("date")),
            pmap.get(p.get("product_id"), "—"),
            float(p.get("quantity", 0)),
            p.get("unit", ""),
            p.get("lot_number", ""),
            float(p.get("cost", 0)),
            p.get("description", ""),
        ])
    buf = _build_xlsx("Üretim", headers, rows, money_cols=(6,), number_cols=(3,))
    return _xlsx_response(buf, "uretim.xlsx")


# ===== DASHBOARD =====
@api.get("/dashboard/summary")
async def dashboard_summary(user: dict = Depends(get_current_user)):
    products = await db.products.find({"active": True}, {"_id": 0}).to_list(5000)
    total_products = len(products)

    def _weight_of(p):
        unit = (p.get("unit") or "").lower()
        stock = float(p.get("current_stock", 0))
        if unit == "kg":
            return stock
        return stock * float(p.get("unit_weight", 0) or 0)

    total_stock_units = sum(float(p.get("current_stock", 0)) for p in products)
    total_stock_weight = sum(_weight_of(p) for p in products)
    total_stock_value_cost = sum(float(p.get("current_stock", 0)) * float(p.get("cost_price", 0)) for p in products)
    total_stock_value_sale = sum(float(p.get("current_stock", 0)) * float(p.get("sale_price", 0)) for p in products)

    today = now_utc().date().isoformat()
    month_start = (now_utc().replace(day=1)).date().isoformat()

    sales_today = await db.sales.find({"date": {"$gte": today, "$lte": today + "T23:59:59"}}, {"_id": 0}).to_list(5000)
    sales_month = await db.sales.find({"date": {"$gte": month_start}}, {"_id": 0}).to_list(5000)

    def _sale_weight(s):
        if "total_weight" in s and s.get("total_weight") is not None:
            return float(s["total_weight"])
        # legacy fallback
        return sum(float(it.get("weight", 0) or 0) for it in s.get("items", []))

    def _sale_qty(s):
        if "total_quantity" in s and s.get("total_quantity") is not None:
            return float(s["total_quantity"])
        return sum(float(it.get("quantity", 0) or 0) for it in s.get("items", []))

    today_total = sum(float(s.get("net_total", 0)) for s in sales_today)
    month_total = sum(float(s.get("net_total", 0)) for s in sales_month)
    today_qty = sum(_sale_qty(s) for s in sales_today)
    today_weight = sum(_sale_weight(s) for s in sales_today)
    month_qty = sum(_sale_qty(s) for s in sales_month)
    month_weight = sum(_sale_weight(s) for s in sales_month)

    # Low stock products
    low_stock = [p for p in products if float(p.get("current_stock", 0)) <= float(p.get("min_stock", 0)) and float(p.get("min_stock", 0)) > 0]

    # Category stock financial value (kept as one chart)
    cat_value = {}
    cat_weight = {}
    cat_sale = {}
    for p in products:
        c = p.get("category") or "Diğer"
        cat_value.setdefault(c, 0.0)
        cat_value[c] += float(p.get("current_stock", 0)) * float(p.get("cost_price", 0))
        cat_weight.setdefault(c, 0.0)
        cat_weight[c] += _weight_of(p)
        cat_sale.setdefault(c, 0.0)
        cat_sale[c] += float(p.get("current_stock", 0)) * float(p.get("sale_price", 0))
    category_value = [{"category": k, "value": v} for k, v in cat_value.items()]
    category_weight = sorted(
        [{"category": k, "weight": v} for k, v in cat_weight.items()],
        key=lambda x: x["weight"], reverse=True,
    )
    category_sale_value = sorted(
        [{"category": k, "value": v} for k, v in cat_sale.items()],
        key=lambda x: x["value"], reverse=True,
    )

    # Top 5 by quantity & weight
    all_sales = await db.sales.find({}, {"_id": 0}).to_list(10000)
    sold_map = {}
    for s in all_sales:
        for it in s.get("items", []):
            key = it["product_id"]
            sold_map.setdefault(key, {"product_id": key, "product_name": it.get("product_name", ""), "quantity": 0.0, "weight": 0.0, "revenue": 0.0})
            sold_map[key]["quantity"] += float(it.get("quantity", 0))
            sold_map[key]["weight"] += float(it.get("weight", 0) or 0)
            sold_map[key]["revenue"] += float(it.get("line_total", 0))
    top5 = sorted(sold_map.values(), key=lambda x: x["quantity"], reverse=True)[:5]

    # Recent productions & sales
    recent_prod = await db.productions.find({}, {"_id": 0}).sort("created_at", -1).to_list(5)
    recent_sales = await db.sales.find({}, {"_id": 0}).sort("created_at", -1).to_list(5)

    # 14-day exit trend (quantity + weight)
    trend = []
    for i in range(13, -1, -1):
        d = (now_utc().date() - timedelta(days=i)).isoformat()
        day_sales = [s for s in all_sales if (s.get("date", "") or "")[:10] == d]
        trend.append({
            "date": d,
            "quantity": sum(_sale_qty(s) for s in day_sales),
            "weight": sum(_sale_weight(s) for s in day_sales),
        })

    return {
        "total_products": total_products,
        "total_stock_units": total_stock_units,
        "total_stock_weight": total_stock_weight,
        "total_stock_value_cost": total_stock_value_cost,
        "total_stock_value_sale": total_stock_value_sale,
        "today_sales": today_total,
        "today_quantity": today_qty,
        "today_weight": today_weight,
        "month_sales": month_total,
        "month_quantity": month_qty,
        "month_weight": month_weight,
        "low_stock_products": low_stock,
        "category_value": category_value,
        "category_weight": category_weight,
        "category_sale_value": category_sale_value,
        "top_products": top5,
        "recent_productions": recent_prod,
        "recent_sales": recent_sales,
        "exit_trend": trend,
    }


# ===== REPORTS =====
@api.get("/reports/stock")
async def report_stock(user: dict = Depends(get_current_user)):
    products = await db.products.find({"active": True}, {"_id": 0}).to_list(5000)
    rows = []
    for p in products:
        stock = float(p.get("current_stock", 0))
        cost = float(p.get("cost_price", 0))
        sale = float(p.get("sale_price", 0))
        unit = (p.get("unit") or "").lower()
        unit_w = float(p.get("unit_weight", 0) or 0)
        stock_weight = stock if unit == "kg" else stock * unit_w
        rows.append({
            "id": p["id"],
            "name": p["name"],
            "code": p["code"],
            "category": p.get("category", ""),
            "unit": p.get("unit", ""),
            "unit_weight": unit_w,
            "current_stock": stock,
            "stock_weight": stock_weight,
            "cost_price": cost,
            "sale_price": sale,
            "stock_value_cost": stock * cost,
            "stock_value_sale": stock * sale,
        })
    return {
        "rows": rows,
        "total_cost_value": sum(r["stock_value_cost"] for r in rows),
        "total_sale_value": sum(r["stock_value_sale"] for r in rows),
        "total_stock_weight": sum(r["stock_weight"] for r in rows),
        "gross_profit_potential": sum(r["stock_value_sale"] - r["stock_value_cost"] for r in rows),
    }


@api.get("/reports/category-distribution")
async def report_category(user: dict = Depends(get_current_user)):
    products = await db.products.find({"active": True}, {"_id": 0}).to_list(5000)
    cats = {}
    for p in products:
        cat = p.get("category") or "Diğer"
        unit = (p.get("unit") or "").lower()
        stock = float(p.get("current_stock", 0))
        weight = stock if unit == "kg" else stock * float(p.get("unit_weight", 0) or 0)
        cats.setdefault(cat, {"category": cat, "units": 0.0, "weight": 0.0, "value": 0.0, "sale_value": 0.0, "count": 0})
        cats[cat]["units"] += stock
        cats[cat]["weight"] += weight
        cats[cat]["value"] += stock * float(p.get("cost_price", 0))
        cats[cat]["sale_value"] += stock * float(p.get("sale_price", 0))
        cats[cat]["count"] += 1
    return sorted(list(cats.values()), key=lambda x: x["weight"], reverse=True)


@api.get("/reports/production")
async def report_production(start: Optional[str] = None, end: Optional[str] = None, user: dict = Depends(get_current_user)):
    q = {}
    if start or end:
        q["date"] = {}
        if start:
            q["date"]["$gte"] = start
        if end:
            q["date"]["$lte"] = end + "T23:59:59"
    prods = await db.productions.find(q, {"_id": 0}).to_list(10000)
    products_all = await db.products.find({}, {"_id": 0}).to_list(5000)
    pmap = {pp["id"]: pp for pp in products_all}

    def _wt(prod_doc, qty):
        p = pmap.get(prod_doc.get("product_id"))
        if not p:
            return 0.0
        unit = (p.get("unit") or "").lower()
        if unit == "kg":
            return float(qty)
        return float(qty) * float(p.get("unit_weight", 0) or 0)

    by_day = {}
    by_product = {}
    for p in prods:
        d = (p.get("date") or "")[:10]
        qty = float(p.get("quantity", 0))
        wt = _wt(p, qty)
        by_day.setdefault(d, {"quantity": 0.0, "weight": 0.0})
        by_day[d]["quantity"] += qty
        by_day[d]["weight"] += wt
        pid = p["product_id"]
        by_product.setdefault(pid, {"product_id": pid, "quantity": 0.0, "weight": 0.0, "count": 0})
        by_product[pid]["quantity"] += qty
        by_product[pid]["weight"] += wt
        by_product[pid]["count"] += 1
    for k in by_product:
        by_product[k]["product_name"] = pmap.get(k, {}).get("name", "Bilinmiyor")
        c = by_product[k]["count"] or 1
        by_product[k]["avg_quantity"] = by_product[k]["quantity"] / c
        by_product[k]["avg_weight"] = by_product[k]["weight"] / c

    total_qty = sum(float(p.get("quantity", 0)) for p in prods)
    total_weight = sum(_wt(p, float(p.get("quantity", 0))) for p in prods)
    days = max(1, len(by_day))
    return {
        "by_day": [{"date": k, "quantity": v["quantity"], "weight": v["weight"]} for k, v in sorted(by_day.items())],
        "by_product": list(by_product.values()),
        "total": total_qty,
        "total_weight": total_weight,
        "daily_avg": total_qty / days,
        "daily_avg_weight": total_weight / days,
        "weekly_avg": total_qty / max(1, days / 7),
        "weekly_avg_weight": total_weight / max(1, days / 7),
        "monthly_avg": total_qty / max(1, days / 30),
        "monthly_avg_weight": total_weight / max(1, days / 30),
    }


@api.get("/reports/sales")
async def report_sales(start: Optional[str] = None, end: Optional[str] = None, user: dict = Depends(get_current_user)):
    q = {}
    if start or end:
        q["date"] = {}
        if start:
            q["date"]["$gte"] = start
        if end:
            q["date"]["$lte"] = end + "T23:59:59"
    sales = await db.sales.find(q, {"_id": 0}).to_list(10000)
    products_all = await db.products.find({}, {"_id": 0}).to_list(5000)
    pmap = {pp["id"]: pp for pp in products_all}

    def _item_weight(it):
        if it.get("weight") is not None:
            return float(it.get("weight", 0) or 0)
        p = pmap.get(it.get("product_id"))
        if not p:
            return 0.0
        unit = (p.get("unit") or "").lower()
        if unit == "kg":
            return float(it.get("quantity", 0))
        return float(it.get("quantity", 0)) * float(p.get("unit_weight", 0) or 0)

    by_day = {}
    by_product = {}
    by_customer = {}
    for s in sales:
        d = (s.get("date") or "")[:10]
        sale_qty = sum(float(it.get("quantity", 0)) for it in s.get("items", []))
        sale_wt = sum(_item_weight(it) for it in s.get("items", []))
        by_day.setdefault(d, {"date": d, "total": 0.0, "quantity": 0.0, "weight": 0.0, "count": 0})
        by_day[d]["total"] += float(s.get("net_total", 0))
        by_day[d]["quantity"] += sale_qty
        by_day[d]["weight"] += sale_wt
        by_day[d]["count"] += 1
        cid = s.get("customer_id")
        by_customer.setdefault(cid, {"customer_id": cid, "customer_name": s.get("customer_name", ""), "revenue": 0.0, "quantity": 0.0, "weight": 0.0, "count": 0})
        by_customer[cid]["revenue"] += float(s.get("net_total", 0))
        by_customer[cid]["quantity"] += sale_qty
        by_customer[cid]["weight"] += sale_wt
        by_customer[cid]["count"] += 1
        for it in s.get("items", []):
            pid = it["product_id"]
            by_product.setdefault(pid, {"product_id": pid, "product_name": it.get("product_name", ""), "quantity": 0.0, "weight": 0.0, "revenue": 0.0, "cost": 0.0})
            by_product[pid]["quantity"] += float(it.get("quantity", 0))
            by_product[pid]["weight"] += _item_weight(it)
            by_product[pid]["revenue"] += float(it.get("line_total", 0))
            by_product[pid]["cost"] += float(it.get("quantity", 0)) * float(it.get("cost_price", 0))
    total = sum(float(s.get("net_total", 0)) for s in sales)
    total_qty = sum(b["quantity"] for b in by_day.values())
    total_weight = sum(b["weight"] for b in by_day.values())
    days = max(1, len(by_day))
    cogs = sum(p["cost"] for p in by_product.values())
    return {
        "by_day": sorted(by_day.values(), key=lambda x: x["date"]),
        "by_product": sorted(by_product.values(), key=lambda x: x["weight"], reverse=True),
        "by_customer": sorted(by_customer.values(), key=lambda x: x["weight"], reverse=True),
        "total": total,
        "total_quantity": total_qty,
        "total_weight": total_weight,
        "count": len(sales),
        "daily_avg": total / days,
        "daily_avg_quantity": total_qty / days,
        "daily_avg_weight": total_weight / days,
        "weekly_avg": total / max(1, days / 7),
        "weekly_avg_weight": total_weight / max(1, days / 7),
        "monthly_avg": total / max(1, days / 30),
        "monthly_avg_weight": total_weight / max(1, days / 30),
        "gross_profit": total - cogs,
        "cogs": cogs,
    }


# ===== USERS (admin) =====
@api.get("/users")
async def list_users(user: dict = Depends(require_roles(ROLE_ADMIN))):
    docs = await db.users.find({}, {"_id": 0, "password_hash": 0}).sort("created_at", -1).to_list(500)
    user_ids = [u["id"] for u in docs]
    pending = await db.password_reset_tokens.find(
        {"user_id": {"$in": user_ids}, "used": False, "admin_request": True},
        {"_id": 0, "user_id": 1, "created_at": 1},
    ).sort("created_at", -1).to_list(1000)
    reset_requests = {}
    for rec in pending:
        reset_requests.setdefault(rec["user_id"], rec.get("created_at"))
    for u in docs:
        if u.get("permissions") is None:
            u["permissions"] = default_permissions(u.get("role", ""))
        if u["id"] in reset_requests:
            u["password_reset_requested"] = True
            u["password_reset_requested_at"] = reset_requests[u["id"]]
        else:
            u["password_reset_requested"] = False
    return docs


@api.post("/users")
async def create_user(payload: UserCreateIn, user: dict = Depends(require_roles(ROLE_ADMIN))):
    email = payload.email.lower().strip()
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=400, detail="Bu e-posta zaten kayıtlı")
    role = payload.role if payload.role in ALL_ROLES else ROLE_STAFF
    new_user = {
        "id": str(uuid.uuid4()),
        "email": email,
        "name": payload.name.strip(),
        "role": role,
        "active": True,
        "language": payload.language or "en",
        "permissions": default_permissions(role),
        "password_hash": hash_password(payload.password),
        "created_at": now_utc().isoformat(),
    }
    await db.users.insert_one(new_user)
    new_user.pop("password_hash", None)
    new_user.pop("_id", None)
    return new_user


@api.put("/users/{uid}")
async def update_user(uid: str, payload: dict, user: dict = Depends(require_roles(ROLE_ADMIN))):
    allowed = {k: v for k, v in payload.items() if k in ["name", "email", "role", "active", "permissions"]}
    password_changed = False
    if "email" in allowed:
        em = (allowed["email"] or "").lower().strip()
        if not em:
            raise HTTPException(status_code=400, detail="E-posta boş olamaz")
        clash = await db.users.find_one({"email": em, "id": {"$ne": uid}})
        if clash:
            raise HTTPException(status_code=400, detail="Bu e-posta başka bir kullanıcıya ait")
        allowed["email"] = em
    if "role" in allowed and allowed["role"] not in ALL_ROLES:
        raise HTTPException(status_code=400, detail="Geçersiz rol")
    if "permissions" in allowed:
        if not isinstance(allowed["permissions"], list):
            raise HTTPException(status_code=400, detail="permissions liste olmalı")
        bad = [p for p in allowed["permissions"] if p not in ALL_PERMISSIONS]
        if bad:
            raise HTTPException(status_code=400, detail=f"Geçersiz izinler: {', '.join(bad)}")
    if "password" in payload and payload["password"]:
        if len(payload["password"]) < 6:
            raise HTTPException(status_code=400, detail="Şifre en az 6 karakter olmalı")
        allowed["password_hash"] = hash_password(payload["password"])
        password_changed = True
    res = await db.users.update_one({"id": uid}, {"$set": allowed})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
    if password_changed:
        await db.password_reset_tokens.update_many(
            {"user_id": uid, "used": False},
            {"$set": {
                "used": True,
                "used_at": now_utc(),
                "request_status": "resolved",
                "resolved_by": user["id"],
            }},
        )
    u = await db.users.find_one({"id": uid}, {"_id": 0, "password_hash": 0})
    if u.get("permissions") is None:
        u["permissions"] = default_permissions(u.get("role", ""))
    pending_request = None if password_changed else await db.password_reset_tokens.find_one(
        {"user_id": uid, "used": False, "admin_request": True},
        {"_id": 0, "created_at": 1},
    )
    u["password_reset_requested"] = bool(pending_request)
    if pending_request:
        u["password_reset_requested_at"] = pending_request.get("created_at")
    return u


@api.delete("/users/{uid}")
async def delete_user(uid: str, user: dict = Depends(require_roles(ROLE_ADMIN))):
    if uid == user["id"]:
        raise HTTPException(status_code=400, detail="Kendi hesabınızı silemezsiniz")
    res = await db.users.delete_one({"id": uid})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
    return {"ok": True}


# ===== Health =====
@api.get("/")
async def root():
    return {"status": "ok", "service": "Stok-Üretim-Satış API"}


# ----- Register router + CORS -----
app.include_router(api)

cors_origins = [origin.strip() for origin in os.environ.get("CORS_ORIGINS", "*").split(",") if origin.strip()]
if not cors_origins:
    cors_origins = ["*"]
cors_origin_regex = os.environ.get("CORS_ORIGIN_REGEX", r"https://.*\.vercel\.app").strip() or None

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=cors_origins,
    allow_origin_regex=cors_origin_regex,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ----- Startup: seed admin + indexes -----
@app.on_event("startup")
async def startup_event():
    try:
        init_storage()
    except Exception as e:
        logger.warning(f"Storage init at startup: {e}")
    try:
        await db.users.create_index("email", unique=True)
        await db.users.create_index("id", unique=True)
        await db.products.create_index("id", unique=True)
        await db.products.create_index("code")
        await db.customers.create_index("id", unique=True)
        await db.productions.create_index("id", unique=True)
        await db.sales.create_index("id", unique=True)
        await db.stock_movements.create_index("id", unique=True)
        await db.company.create_index("id", unique=True)
        await db.password_reset_tokens.create_index("user_id")
    except Exception as e:
        logger.warning(f"Index creation: {e}")

    admin_email = os.environ.get("ADMIN_EMAIL", "admin@stoktakip.com").lower()
    admin_pw = os.environ.get("ADMIN_PASSWORD", "Admin123!")
    existing = await db.users.find_one({"email": admin_email})
    if existing is None:
        await db.users.insert_one({
            "id": str(uuid.uuid4()),
            "email": admin_email,
            "name": "Admin",
            "role": ROLE_ADMIN,
            "active": True,
            "language": "en",
            "permissions": list(ALL_PERMISSIONS),
            "password_hash": hash_password(admin_pw),
            "created_at": now_utc().isoformat(),
        })
        logger.info(f"Admin user seeded: {admin_email}")
    elif not verify_password(admin_pw, existing.get("password_hash", "")):
        await db.users.update_one({"email": admin_email}, {"$set": {"password_hash": hash_password(admin_pw), "role": ROLE_ADMIN, "active": True}})
        logger.info(f"Admin password refreshed: {admin_email}")


@app.on_event("shutdown")
async def shutdown_event():
    client.close()
